import json
import openpyxl





work_book = 'e.xlsx'
wb = openpyxl.load_workbook(work_book)
ws = wb["List1"]
cities_name_en = "gg"

f = open("districts_lite.json", encoding="utf-8")
districts = json.load(f)
f.close()
f = open("regions_lite.json", encoding="utf-8")
regions = json.load(f)
f.close()
f = open("cities_lite.json", encoding="utf-8")
cities = json.load(f)
f.close()
prefix = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T']
wregion_id = ""
wcity_id =""
column = 0
row = 0
for item in districts:
    region_id = item.get('region_id')

    if region_id != wregion_id:
        for regions_item in regions:

            if regions_item.get('region_id') == region_id:
                regions_name_en = regions_item['name_en']
                print(region_id, regions_name_en)

                wb.create_sheet(regions_name_en)
                ws = wb[regions_name_en]
                column = 0
                row = 1
                #wb.save(work_book)

        wregion_id = region_id

    city_id = item.get('city_id')
    if city_id != wcity_id:
        for cities_item in cities:

            if cities_item.get('city_id') == city_id:
                cities_name_en = cities_item['name_en']
                print(region_id,city_id, cities_name_en)
        column +=1
        row = 1
        ws.cell(row = row, column= column, value = cities_name_en)
        row +=1
        wcity_id = city_id
    name_en = item.get('name_en')

    ws.cell(row = row, column= column, value = name_en)
    row += 1

    print(region_id, city_id, name_en)
    
    wb.save(work_book)
    #name = item.get('name')
    #for s in x.get('structure'):
    #    #write this string to a csv file
    #    print(name + ' ' + s.get('column') + ' ' + s.get('type'))
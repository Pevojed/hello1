import openpyxl
from datetime import datetime

# current date and time
date_time = datetime.now()



sheets = ['NORAM', 'LATAM', 'EMEA', 'APAC']
first_line = f"Region, FBPN, Type, Month, Year, Planned_Orders, Notes"
first_row = ['Region', 'FBPN', 'Type', 'Month', 'Year', 'Planned_Orders', 'Notes']



work_book = 'Forecast_30NOV23.xlsx'
try:
    wb = openpyxl.load_workbook(work_book, data_only=True)
except FileNotFoundError:
    print(f"No such file or directory {work_book} ")
except:
    print("Something went wrong")
naw=date_time.strftime('%d%B%Y')
naw= naw[0:2]+naw[2:5].upper() +naw[-2:]

work_book1 = 'Forecast_30NOV23'+'consolidated_' +naw + '.xlsx'

wb1=openpyxl.load_workbook('Forecast_30NOV23.xlsx', data_only=True)

wb1.save(work_book1)
wb1.create_sheet('Data')

ws1 = wb1['Data']
for i in range(0, 7):
    ws1.cell(row=1, column=i+1).value=first_row[i]
ws1_row =1


file = open('log.txt', 'w')
file.write(first_line + "\n")
for sheet in sheets:
    ws=wb[sheet]
    print("***", sheet, "***")
    last_row = len(ws['A'])
    print(last_row)
    for i in range(3,last_row):
        region = ''
        fbpn = ''
        type = ''
        notes = ''
        if (((str(ws.cell(row =i, column =1).value) !='') and (ws.cell(row =i, column =1).value) != None)):

            region = str(ws.cell(row = i,column = 1).value)
            fbpn = str(ws.cell(row = i,column = 2).value)
            type = str(ws.cell(row = i,column = 3).value)
            if (((str(ws.cell(row=i, column=22).value) != '') and (ws.cell(row=i, column=22).value) != None)):
                notes = str(ws.cell(row = i,column = 22).value)
            for k in range(5, 21):
                month = ''
                planed_orders = ''
                year =''
                if (str(ws.cell(row = (i+2), column =k ).value) != "0"):
                    month = str(ws.cell(row = 2, column =k ).value)
                    planned_orders = int(ws.cell(row = i+2, column =k ).value)
                    for m in  range(k,4,-1):

                        if str(ws.cell(row = 1, column =m ).value)[0] =="2":
                            year = int(ws.cell(row = 1, column =m ).value)
                            break
                    line = f"{region},{fbpn},{type},{month},{year},{planned_orders},{notes}"
                    print(line)
                    file.write(line + "\n")
                    #write in .xlsx
                    ws1_row +=1

                    ws1.cell(row=ws1_row, column=1).value = region
                    ws1.cell(row=ws1_row, column=2).value = fbpn
                    ws1.cell(row=ws1_row, column=3).value = type
                    ws1.cell(row=ws1_row, column=4).value = month
                    ws1.cell(row=ws1_row, column=5).value = year
                    ws1.cell(row=ws1_row, column=6).value = planned_orders
                    if ((notes != '') and (notes != None)):
                        ws1.cell(row=ws1_row, column=7).value = notes


                    #lines.append(line)

wb1.save(work_book1)
wb.close()
wb.close()
file.close()

